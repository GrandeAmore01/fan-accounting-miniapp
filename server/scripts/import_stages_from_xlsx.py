# -*- coding: utf-8 -*-
"""从 data_s/时代少年团小程序公共数据库完整版.xlsx 生成舞台相关 MySQL 种子数据。"""

import os
import re
import json
import openpyxl

ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..'))
DATA_DIR = os.path.join(ROOT, 'data_s')
XLSX_NAME = '时代少年团小程序公共数据库完整版.xlsx'
OUTPUT_SQL = os.path.join(ROOT, 'server', 'sql', 'stages_data.sql')


def get_sheet_rows(workbook, part):
    for sheet_name in workbook.sheetnames:
        if part in sheet_name:
            rows = []
            for row in workbook[sheet_name].iter_rows(values_only=True):
                rows.append([str(cell).strip() if cell is not None else '' for cell in row])
            return sheet_name, rows
    return None, []


def sql_escape(value):
    if value is None:
        return ''
    return str(value).replace('\\', '\\\\').replace("'", "''")


def normalize_date(value):
    text = (value or '').strip()
    if not text:
        return ''
    text = text.replace('/', '-').replace('.', '-')
    parts = [part for part in text.split('-') if part]
    if len(parts) != 3:
        return text
    year, month, day = parts
    return f'{year.zfill(4)}-{month.zfill(2)}-{day.zfill(2)}'


def album_id_from_name(name, index):
    return f'album_{index:03d}'


def detect_stage_type(tour_name, city):
    text = f'{tour_name}{city}'
    if '音乐节' in text or '拼盘' in text:
        return 'festival'
    return 'concert'


def detect_is_online(city, tour_name):
    text = f'{city}{tour_name}'
    return 1 if '线上' in text else 0


def parse_price(value):
    text = str(value or '').strip()
    if not text or text in {'无', '暂无', '-', 'None'}:
        return None
    match = re.search(r'\d+(?:\.\d+)?', text)
    if not match:
        return None
    return int(float(match.group()))


def build_stage_name(tour_name, city):
    tour = tour_name.strip()
    city_name = city.strip()
    if city_name and city_name not in tour:
        return f'{tour}·{city_name}站'
    return tour


def main():
    xlsx_path = os.path.join(DATA_DIR, XLSX_NAME)
    if not os.path.exists(xlsx_path):
        raise FileNotFoundError(f'未找到数据文件: {xlsx_path}')

    workbook = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

    _, song_rows = get_sheet_rows(workbook, 'song_')
    _, stage_rows = get_sheet_rows(workbook, 'stage_')
    _, link_rows = get_sheet_rows(workbook, 'song-stage')
    _, price_rows = get_sheet_rows(workbook, 'price-stage')
    workbook.close()

    song_map = {}
    album_map = {}
    album_name_to_id = {}
    for row in song_rows[1:]:
        if not row or not row[0]:
            continue
        song_id, album_name, release_date, song_name = row[0], row[1], row[2], row[3]
        if album_name and album_name not in album_name_to_id:
            album_name_to_id[album_name] = f'album_{len(album_name_to_id) + 1:03d}'
        album_id = album_name_to_id.get(album_name, '')
        song_map[song_id] = {
            'song_id': song_id,
            'album_id': album_id,
            'album_name': album_name,
            'release_date': release_date,
            'song_name': song_name
        }
        if album_name and album_id not in album_map:
            year_match = re.search(r'(20\d{2})', release_date or '')
            album_map[album_id] = {
                'album_id': album_id,
                'album_name': album_name,
                'album_name_cn': album_name,
                'release_year': int(year_match.group(1)) if year_match else None,
                'sort_order': len(album_map) + 1
            }

    stage_song_links = []
    stage_songs = {}
    for row in link_rows[1:]:
        if len(row) < 2 or not row[0] or not row[1]:
            continue
        stage_id, song_id = row[0], row[1]
        stage_song_links.append((stage_id, song_id))
        stage_songs.setdefault(stage_id, []).append(song_id)

    price_map = {}
    for row in price_rows[1:]:
        if len(row) < 2 or not row[0]:
            continue
        stage_id = row[0]
        price = parse_price(row[1])
        if price is None:
            continue
        price_map.setdefault(stage_id, [])
        if price not in price_map[stage_id]:
            price_map[stage_id].append(price)

    for stage_id in price_map:
        price_map[stage_id].sort()

    stage_records = []
    for index, row in enumerate(stage_rows[1:], start=1):
        if not row or not row[0]:
            continue
        stage_id, tour_name, city, stage_date, venue = row[0], row[1], row[2], row[3], row[4]
        date_text = normalize_date(stage_date)
        year = int(date_text.split('-')[0]) if date_text else 0
        linked_songs = stage_songs.get(stage_id, [])
        album_ids = []
        for song_id in linked_songs:
            song = song_map.get(song_id)
            if song and song['album_id'] not in album_ids:
                album_ids.append(song['album_id'])
        album_id = album_ids[0] if album_ids else ''
        price_tiers = price_map.get(stage_id, [])
        ticket_price = price_tiers[0] if price_tiers else 0
        stage_records.append({
            'stage_id': stage_id,
            'stage_name': build_stage_name(tour_name, city),
            'stage_type': detect_stage_type(tour_name, city),
            'year': year,
            'stage_date': date_text,
            'city': city,
            'venue': venue,
            'location': city,
            'album_id': album_id,
            'price_tiers': price_tiers,
            'is_online': detect_is_online(city, tour_name),
            'ticket_price': ticket_price,
            'description': tour_name,
            'sort_order': index
        })

    lines = [
        'USE fan_accounting;',
        '',
        'CREATE TABLE IF NOT EXISTS albums (',
        "  album_id VARCHAR(64) PRIMARY KEY,",
        "  album_name VARCHAR(255) NOT NULL,",
        "  album_name_cn VARCHAR(255) NOT NULL DEFAULT '',",
        '  release_year INT NULL,',
        '  sort_order INT NOT NULL DEFAULT 0',
        ") ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci;",
        '',
        'CREATE TABLE IF NOT EXISTS songs (',
        "  song_id VARCHAR(64) PRIMARY KEY,",
        "  song_name VARCHAR(255) NOT NULL,",
        "  album_id VARCHAR(64) NOT NULL DEFAULT '',",
        '  sort_order INT NOT NULL DEFAULT 0,',
        '  INDEX idx_songs_album (album_id)',
        ") ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci;",
        '',
        'CREATE TABLE IF NOT EXISTS stages (',
        "  stage_id VARCHAR(64) PRIMARY KEY,",
        "  stage_name VARCHAR(255) NOT NULL,",
        "  stage_type VARCHAR(32) NOT NULL DEFAULT 'concert',",
        '  year INT NOT NULL,',
        '  stage_date DATE NOT NULL,',
        "  city VARCHAR(120) NOT NULL DEFAULT '',",
        "  venue VARCHAR(255) NOT NULL DEFAULT '',",
        "  location VARCHAR(255) NOT NULL DEFAULT '',",
        "  album_id VARCHAR(64) NOT NULL DEFAULT '',",
        '  price_tiers_json JSON NULL,',
        '  is_online TINYINT(1) NOT NULL DEFAULT 0,',
        '  ticket_price DECIMAL(10,2) NOT NULL DEFAULT 0.00,',
        "  description VARCHAR(1000) NOT NULL DEFAULT '',",
        '  sort_order INT NOT NULL DEFAULT 0,',
        '  INDEX idx_stages_year (year),',
        '  INDEX idx_stages_type (stage_type),',
        '  INDEX idx_stages_date (stage_date)',
        ") ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci;",
        '',
        'CREATE TABLE IF NOT EXISTS stage_songs (',
        "  stage_id VARCHAR(64) NOT NULL,",
        "  song_id VARCHAR(64) NOT NULL,",
        '  sort_order INT NOT NULL DEFAULT 0,',
        '  PRIMARY KEY (stage_id, song_id),',
        '  INDEX idx_stage_songs_song (song_id)',
        ") ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci;",
        '',
        'DELETE FROM stage_songs;',
        'DELETE FROM stages;',
        'DELETE FROM songs;',
        'DELETE FROM albums;',
        ''
    ]

    lines.append('INSERT INTO albums (album_id, album_name, album_name_cn, release_year, sort_order) VALUES')
    album_values = []
    for album in album_map.values():
        release_year = 'NULL' if album['release_year'] is None else str(album['release_year'])
        album_values.append(
            f"('{sql_escape(album['album_id'])}', '{sql_escape(album['album_name'])}', "
            f"'{sql_escape(album['album_name_cn'])}', {release_year}, {album['sort_order']})"
        )
    lines.append(',\n'.join(album_values) + ';')
    lines.append('')

    lines.append('INSERT INTO songs (song_id, song_name, album_id, sort_order) VALUES')
    song_values = []
    for index, song in enumerate(song_map.values(), start=1):
        song_values.append(
            f"('{sql_escape(song['song_id'])}', '{sql_escape(song['song_name'])}', "
            f"'{sql_escape(song['album_id'])}', {index})"
        )
    lines.append(',\n'.join(song_values) + ';')
    lines.append('')

    lines.append(
        'INSERT INTO stages (stage_id, stage_name, stage_type, year, stage_date, city, venue, location, '
        'album_id, price_tiers_json, is_online, ticket_price, description, sort_order) VALUES'
    )
    stage_values = []
    for stage in stage_records:
        price_json = json.dumps(stage['price_tiers'], ensure_ascii=False)
        stage_values.append(
            f"('{sql_escape(stage['stage_id'])}', '{sql_escape(stage['stage_name'])}', "
            f"'{sql_escape(stage['stage_type'])}', {stage['year']}, '{sql_escape(stage['stage_date'])}', "
            f"'{sql_escape(stage['city'])}', '{sql_escape(stage['venue'])}', '{sql_escape(stage['location'])}', "
            f"'{sql_escape(stage['album_id'])}', '{sql_escape(price_json)}', {stage['is_online']}, "
            f"{stage['ticket_price']}, '{sql_escape(stage['description'])}', {stage['sort_order']})"
        )
    lines.append(',\n'.join(stage_values) + ';')
    lines.append('')

    lines.append('INSERT INTO stage_songs (stage_id, song_id, sort_order) VALUES')
    link_values = []
    stage_link_order = {}
    for stage_id, song_id in stage_song_links:
        stage_link_order[stage_id] = stage_link_order.get(stage_id, 0) + 1
        link_values.append(
            f"('{sql_escape(stage_id)}', '{sql_escape(song_id)}', {stage_link_order[stage_id]})"
        )
    lines.append(',\n'.join(link_values) + ';')
    lines.append('')

    with open(OUTPUT_SQL, 'w', encoding='utf-8') as handle:
        handle.write('\n'.join(lines))

    summary = {
        'source': XLSX_NAME,
        'albums': len(album_map),
        'songs': len(song_map),
        'stages': len(stage_records),
        'stage_songs': len(stage_song_links),
        'output': OUTPUT_SQL
    }
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == '__main__':
    main()
