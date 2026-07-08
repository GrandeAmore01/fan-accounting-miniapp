"""从公共数据库 Excel 生成藏品 MySQL 数据脚本。"""

import os
import re
from datetime import date, datetime
from decimal import Decimal, InvalidOperation

import openpyxl


ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..'))
XLSX_PATH = os.path.join(
    ROOT,
    'data_s',
    '时代少年团小程序公共数据库完整版.xlsx'
)
OUTPUT_PATH = os.path.join(
    ROOT,
    'server',
    'sql',
    'collections_data.sql'
)

CATEGORY_MAP = {
    '周边': 'goods',
    '杂志': 'magazine',
    '应援服': 'support_wear',
    '伴手礼': 'gift'
}


def find_collection_sheet(workbook):
    """寻找藏品图鉴工作表。"""
    for sheet_name in workbook.sheetnames:
        if 'collection_' in sheet_name:
            return workbook[sheet_name]
    raise ValueError('未找到 collection_藏品图鉴表')


def text_value(value):
    """保留 Excel 中的原始文字。"""
    if value is None:
        return ''

    if isinstance(value, datetime):
        return value.strftime('%Y-%m-%d')

    if isinstance(value, date):
        return value.strftime('%Y-%m-%d')

    if isinstance(value, float):
        return format(value, 'g')

    return str(value).strip()


def parse_date(value):
    """
    仅将完整的年月日转换为 MySQL DATE。
    只有年月或无法确认具体日期时返回 None。
    """
    if isinstance(value, datetime):
        return value.date().isoformat()

    if isinstance(value, date):
        return value.isoformat()

    if not isinstance(value, str):
        return None

    value = value.strip()
    matched = re.fullmatch(
        r'(\d{4})[./-](\d{1,2})[./-](\d{1,2})',
        value
    )

    if not matched:
        return None

    year, month, day = map(int, matched.groups())

    try:
        return date(year, month, day).isoformat()
    except ValueError:
        return None


def parse_price(value):
    """
    同时保存原始价格文字和可计算的数字价格。
    不修改 Excel 中的原始含义。
    """
    price_text = text_value(value)

    if re.fullmatch(r'\d+(?:\.\d{1,2})?[rR]?', price_text):
        number_text = price_text.rstrip('rR')

        try:
            price = Decimal(number_text).quantize(Decimal('0.01'))
            return price, price_text, 'purchase', ''
        except InvalidOperation:
            pass

    if price_text == '门票':
        return (
            None,
            price_text,
            'ticket_gift',
            '凭对应场次门票获得，无单独数字售价'
        )

    return (
        None,
        price_text,
        'special',
        f'原始价格：{price_text}'
    )


def normalize_stage_id(value):
    """将“无”转换为空关联。"""
    stage_id = text_value(value)

    if not stage_id or stage_id == '无':
        return None

    return stage_id


def sql_text(value):
    """生成安全的 SQL 字符串。"""
    if value is None:
        return 'NULL'

    escaped = str(value).replace('\\', '\\\\').replace("'", "''")
    return f"'{escaped}'"


def sql_decimal(value):
    if value is None:
        return 'NULL'
    return f'{value:.2f}'


def build_record(headers, row):
    """把一行 Excel 数据转换成统一结构。"""
    source = dict(zip(headers, row))

    collection_id = text_value(source.get('collection_id'))
    collection_name = text_value(source.get('藏品全称'))
    source_category = text_value(source.get('藏品分类'))

    reference_price, price_text, acquisition_type, price_note = (
        parse_price(source.get('售价'))
    )

    return {
        'collection_id': collection_id,
        'collection_name': collection_name,
        'sale_type': text_value(source.get('贩卖类别')),
        'category': CATEGORY_MAP.get(source_category, 'other'),
        'sale_date': parse_date(
            source.get('开售时间/购买时间/领取时间')
        ),
        'sale_date_text': text_value(
            source.get('开售时间/购买时间/领取时间')
        ),
        'stage_id': normalize_stage_id(source.get('对应舞台')),
        'reference_price': reference_price,
        'price_text': price_text,
        'acquisition_type': acquisition_type,
        'price_note': price_note,
        'brand': text_value(source.get('代言品牌')),
        'series_name': text_value(source.get('所属系列')),
        'image_url': text_value(source.get('商品图url')),
        'source_category': source_category
    }


def create_table_sql():
    return """CREATE TABLE IF NOT EXISTS collections (
  collection_id VARCHAR(64) PRIMARY KEY,
  collection_name VARCHAR(255) NOT NULL,
  sale_type VARCHAR(64) NOT NULL DEFAULT '',
  category VARCHAR(32) NOT NULL DEFAULT 'goods',
  sale_date DATE NULL,
  sale_date_text VARCHAR(64) NOT NULL DEFAULT '',
  stage_id VARCHAR(64) NULL,
  reference_price DECIMAL(10,2) NULL,
  price_text VARCHAR(64) NOT NULL DEFAULT '',
  acquisition_type VARCHAR(32) NOT NULL DEFAULT 'purchase',
  price_note VARCHAR(255) NOT NULL DEFAULT '',
  brand VARCHAR(255) NOT NULL DEFAULT '',
  series_name VARCHAR(255) NOT NULL DEFAULT '',
  image_url VARCHAR(500) NOT NULL DEFAULT '',
  created_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP,
  updated_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP
    ON UPDATE CURRENT_TIMESTAMP,
  INDEX idx_collections_name (collection_name),
  INDEX idx_collections_category (category),
  INDEX idx_collections_stage (stage_id)
) ENGINE=InnoDB
  DEFAULT CHARSET=utf8mb4
  COLLATE=utf8mb4_unicode_ci;"""


def record_to_sql(record):
    values = [
        sql_text(record['collection_id']),
        sql_text(record['collection_name']),
        sql_text(record['sale_type']),
        sql_text(record['category']),
        sql_text(record['sale_date']),
        sql_text(record['sale_date_text']),
        sql_text(record['stage_id']),
        sql_decimal(record['reference_price']),
        sql_text(record['price_text']),
        sql_text(record['acquisition_type']),
        sql_text(record['price_note']),
        sql_text(record['brand']),
        sql_text(record['series_name']),
        sql_text(record['image_url'])
    ]

    return '(' + ', '.join(values) + ')'


def main():
    if not os.path.exists(XLSX_PATH):
        raise FileNotFoundError(f'未找到 Excel 文件：{XLSX_PATH}')

    workbook = openpyxl.load_workbook(
        XLSX_PATH,
        read_only=True,
        data_only=True
    )

    sheet = find_collection_sheet(workbook)
    rows = list(sheet.iter_rows(values_only=True))
    workbook.close()

    if not rows:
        raise ValueError('藏品工作表为空')

    headers = [text_value(value) for value in rows[0]]
    records = []

    for row_number, row in enumerate(rows[1:], start=2):
        if not any(value is not None and value != '' for value in row):
            continue

        record = build_record(headers, row)

        if not record['collection_id']:
            raise ValueError(f'第 {row_number} 行缺少 collection_id')

        if not record['collection_name']:
            raise ValueError(f'第 {row_number} 行缺少藏品名称')

        records.append(record)

    collection_ids = [item['collection_id'] for item in records]
    duplicate_ids = sorted({
        item_id
        for item_id in collection_ids
        if collection_ids.count(item_id) > 1
    })

    if duplicate_ids:
        raise ValueError(
            f'发现重复 collection_id：{", ".join(duplicate_ids)}'
        )

    unknown_categories = sorted({
        item['source_category']
        for item in records
        if item['source_category'] not in CATEGORY_MAP
    })

    sql_lines = [
        'USE fan_accounting;',
        '',
        create_table_sql(),
        '',
        'DELETE FROM collections;',
        '',
        (
            'INSERT INTO collections ('
            'collection_id, collection_name, sale_type, category, '
            'sale_date, sale_date_text, stage_id, reference_price, '
            'price_text, acquisition_type, price_note, brand, '
            'series_name, image_url'
            ') VALUES'
        ),
        ',\n'.join(record_to_sql(record) for record in records) + ';',
        ''
    ]

    with open(OUTPUT_PATH, 'w', encoding='utf-8', newline='\n') as output:
        output.write('\n'.join(sql_lines))

    numeric_price_count = sum(
        item['reference_price'] is not None
        for item in records
    )
    ticket_gift_count = sum(
        item['acquisition_type'] == 'ticket_gift'
        for item in records
    )
    special_price_count = sum(
        item['acquisition_type'] == 'special'
        for item in records
    )

    print(f'Excel: {XLSX_PATH}')
    print(f'工作表: {sheet.title}')
    print(f'collections: {len(records)}')
    print(f'duplicate_ids: {len(duplicate_ids)}')
    print(f'numeric_prices: {numeric_price_count}')
    print(f'ticket_gifts: {ticket_gift_count}')
    print(f'special_prices: {special_price_count}')
    print(f'unknown_categories: {unknown_categories}')
    print(f'output: {OUTPUT_PATH}')


if __name__ == '__main__':
    main()